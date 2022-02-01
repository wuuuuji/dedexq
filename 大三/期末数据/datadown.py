import requests
from bs4 import BeautifulSoup
import openpyxl


def datadown():
    headers = {
        'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36",
        'Accept-Encoding': "gzip, deflate, br",
        'Accept': "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
    }
    n = int(input('输入需要多少期: '))
    url = "https://chart.cp.360.cn/kaijiang/ssq?lotId=22005a1&chartType=undefined&spanType=0&span={}".format(n)
    respons = requests.get(url, headers=headers)
    # 服务器返回的所有信息
    text = respons.text
    # 要response中现成的字符串
    soup = BeautifulSoup(text, 'lxml')

    tbody = soup.find_all('thead', class_="kaijiang")[0]        # 找到HTML中 thead标签，class=kaijiang的地方
    tbody_th = tbody.find_all('th')     # thead标签下，所有<th>标签。内容以列表的形势存的[ <th width="6%">注数</th>, <th width="9%">奖金（元）</th>]
    # 获取标题

    wb = openpyxl.Workbook()  # 新建excel文件
    ws = wb.active  # 激活sheet，用于后续将数据写入
    ws.title = '双色球中奖信息'  # 指定excel的名称

    ws.cell(row=1, column=1, value=list(tbody_th[0].stripped_strings)[0])  # 期号
    # cell 方法给excel写入数据，row= 行，column=列，value=要写入的值
    ws.cell(row=1, column=2, value=list(tbody_th[1].stripped_strings)[0])  # 开奖日期
    ws.cell(row=1, column=3, value=list(tbody_th[-6].stripped_strings)[0])  # 红球
    ws.cell(row=1, column=4, value=list(tbody_th[-5].stripped_strings)[0])  # 蓝球

    tbody = soup.find_all('tbody', id="data-tab")[0]        # 找到HTML中 tbody标签，class=data-tab的地方
    trs = tbody.find_all('tr')      # tbody标签下，所有<tr>标签。
    # 获取数据

    data_list = []  # 要写入excel的数据
    red_list = []  # 红球
    blue_list = []  # 蓝球
    for tr in trs:
        tds = tr.find_all('td')[:4]
        blue_list.append(list(tds[3].stripped_strings)[0])  # 获取蓝球号码
        tds_text = []  # 中奖号码信息
        redBall = ''
        for index, td in enumerate(tds):
            if index == 2:  # 红球
                for i in list(td.stripped_strings):
                    redBall = redBall + ' ' + i
                    red_list.append(i)
                tds_text.append(redBall.lstrip())
            else:
                tds_text.append(list(td.stripped_strings)[0])
        data_list.append(tds_text)
    for i, dl in enumerate(data_list):
        print("写入第{}行数据".format(i+1))
        for j, dt in enumerate(dl):
            ws.cell(row=i + 2, column=j + 1, value=dt)  # 将中奖号码信息写入excel中
    wb.save('双色球中奖信息.xlsx')  # 将数据保存到本地excel中
    print("已生成xlsx文档")
    return red_list, blue_list

datadown()